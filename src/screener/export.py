"""
src/screener/export.py
======================
Day 17 — Generate output/screener_output.xlsx.

One sheet per preset (6 sheets), sorted by screener_composite_score descending.
20 KPI columns per row. Cells colour-coded green (meets threshold) / red (fails).

Winsorization is computed ONCE in engine init (via composite_score.py) and reused
across all 6 sheets — cross-preset composite scores are comparable.

Author: Bluestock Data Analytics Team
Sprint: 3 — Day 17
"""

from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl.styles import PatternFill, Font, Alignment

from src.screener.engine import FilterEngine
from src.screener.presets import run_all_presets

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Colour fills
# ---------------------------------------------------------------------------
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# ---------------------------------------------------------------------------
# 20 KPI columns (spec-exact order)
# ---------------------------------------------------------------------------
KPI_COLUMNS = [
    "company_id",
    "broad_sector",
    "year",
    "return_on_equity_pct",
    "return_on_capital_employed_pct",
    "net_profit_margin_pct",
    "debt_to_equity",
    "interest_coverage",
    "free_cash_flow_cr",
    "revenue_cagr_5yr",
    "pat_cagr_5yr",
    "sales",
    "net_profit",
    "market_cap_crore",
    "pe_ratio",
    "pb_ratio",
    "dividend_yield_pct",
    "dividend_payout_ratio_pct",
    "screener_composite_score",
    "sector_relative_score",
]

# ---------------------------------------------------------------------------
# Preset thresholds for colour-coding
# Maps preset_name → { column_name: (operator, threshold) }
# operator: "min" = cell passes if value >= threshold
#           "max" = cell passes if value <= threshold
# ---------------------------------------------------------------------------
PRESET_THRESHOLDS: Dict[str, Dict[str, tuple]] = {
    "Quality Compounder": {
        "return_on_equity_pct": ("min", 15.0),
        "debt_to_equity": ("max", 1.0),
        "free_cash_flow_cr": ("min", 0.0),
        "revenue_cagr_5yr": ("min", 10.0),
    },
    "Value Pick": {
        "pe_ratio": ("max", 25.0),
        "pb_ratio": ("max", 5.0),
        "debt_to_equity": ("max", 2.0),
        "dividend_yield_pct": ("min", 0.5),
    },
    "Growth Accelerator": {
        "pat_cagr_5yr": ("min", 20.0),
        "revenue_cagr_5yr": ("min", 15.0),
        "debt_to_equity": ("max", 2.0),
    },
    "Dividend Champion": {
        "dividend_yield_pct": ("min", 2.0),
        "dividend_payout_ratio_pct": ("max", 80.0),
        "free_cash_flow_cr": ("min", 0.0),
    },
    "Debt-Free Blue Chip": {
        "debt_to_equity": ("max", 0.1),
        "return_on_equity_pct": ("min", 12.0),
        "sales": ("min", 5000.0),
    },
    "Turnaround Watch": {
        "free_cash_flow_cr": ("min", 0.0),
    },
}


def _cell_passes_threshold(value: Any, operator: str, threshold: float) -> bool:
    """Check if a cell value meets the preset threshold."""
    if value is None or pd.isna(value):
        return False
    try:
        val = float(value)
    except (TypeError, ValueError):
        return False
    if operator == "min":
        return val >= threshold
    elif operator == "max":
        return val <= threshold
    return False


def export_screener_output(
    engine: FilterEngine,
    output_path: str = "output/screener_output.xlsx",
) -> str:
    """
    Generate screener_output.xlsx with 6 sheets (one per preset).

    Each sheet has 20 KPI columns, sorted by screener_composite_score descending.
    Cells are colour-coded green/red based on preset-specific thresholds.

    Args:
        engine:      Initialized FilterEngine with composite scores computed.
        output_path: Path to write the xlsx file.

    Returns:
        The absolute path to the generated file.
    """
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    results = run_all_presets(engine)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for preset_name, df in results.items():
            # Select and order the 20 KPI columns
            available_cols = [c for c in KPI_COLUMNS if c in df.columns]
            sheet_df = df[available_cols].copy()

            # Truncate sheet name to 31 chars (Excel limit)
            sheet_name = preset_name[:31]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Access the openpyxl worksheet for formatting
            ws = writer.sheets[sheet_name]

            # --- Header formatting ---
            for col_idx in range(1, len(available_cols) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            # --- Column width auto-sizing ---
            for col_idx, col_name in enumerate(available_cols, 1):
                max_len = max(
                    len(str(col_name)),
                    max(
                        (
                            len(str(ws.cell(row=r, column=col_idx).value or ""))
                            for r in range(2, ws.max_row + 1)
                        ),
                        default=0,
                    ),
                )
                ws.column_dimensions[
                    ws.cell(row=1, column=col_idx).column_letter
                ].width = min(max_len + 3, 25)

            # --- Number formatting for financial columns ---
            for col_idx, col_name in enumerate(available_cols, 1):
                if col_name in ("screener_composite_score", "sector_relative_score"):
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = "0.00"
                elif "pct" in col_name or "ratio" in col_name or "cagr" in col_name:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = "0.00"
                elif col_name in (
                    "sales",
                    "net_profit",
                    "market_cap_crore",
                    "free_cash_flow_cr",
                    "interest_coverage",
                    "enterprise_value_crore",
                ):
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = "#,##0.00"

            # --- Conditional colour-coding ---
            thresholds = PRESET_THRESHOLDS.get(preset_name, {})
            col_name_to_idx = {name: idx + 1 for idx, name in enumerate(available_cols)}

            for col_name, (operator, threshold) in thresholds.items():
                if col_name not in col_name_to_idx:
                    continue
                col_idx = col_name_to_idx[col_name]

                for row in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    if _cell_passes_threshold(cell.value, operator, threshold):
                        cell.fill = GREEN_FILL
                    else:
                        cell.fill = RED_FILL

            log.info(
                "  Sheet '%s': %d companies, %d columns",
                sheet_name,
                len(sheet_df),
                len(available_cols),
            )

    abs_path = os.path.abspath(output_path)
    log.info("screener_output.xlsx generated: %s", abs_path)
    return abs_path


# ---------------------------------------------------------------------------
# CLI entry-point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import logging as _logging

    _logging.basicConfig(
        level=_logging.INFO,
        format="%(asctime)s | %(levelname)-8s | %(message)s",
    )

    eng = FilterEngine(
        db_path=Path("db/nifty100.db"),
        config_path=Path("config/screener_config.yaml"),
    )

    output = export_screener_output(eng)
    print(f"\nExport complete: {output}")
