"""
src/analytics/export_peers.py
=============================
Day 20 — Peer Comparison Excel Export.

Generates `output/peer_comparison.xlsx` containing one sheet per peer group.
Outputs the 20-column KPI set from Day 17 (plus company_name and the 2 missing core metrics).
Applies a discrete 3-band color scale to the 10 core metrics based on `percentile_rank`:
- Green: >= 0.75
- Yellow: 0.25 - 0.75
- Red: <= 0.25

Features:
- Benchmarks are highlighted in Gold (only identifying columns).
- Summary median row per peer group.
- year_mismatch_flag annotations (*).
"""

import logging
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from src.screener.engine import FilterEngine
from src.screener.export import KPI_COLUMNS
from src.analytics.peer import METRICS as RANKED_METRICS

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s | %(levelname)-8s | %(message)s"
)
logger = logging.getLogger(__name__)

# Fills
GREEN_FILL = PatternFill(start_color="63BE7B", end_color="63BE7B", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB84", end_color="FFEB84", fill_type="solid")
RED_FILL = PatternFill(start_color="F8696B", end_color="F8696B", fill_type="solid")
GOLD_FILL = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)


def get_percentile_fill(rank: float) -> PatternFill | None:
    if pd.isna(rank):
        return None
    if rank >= 0.75:
        return GREEN_FILL
    elif rank <= 0.25:
        return RED_FILL
    else:
        return YELLOW_FILL


def generate_peer_comparison(
    db_path: Path = Path("db/nifty100.db"),
    out_path: Path = Path("output/peer_comparison.xlsx"),
):
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # 1. Initialize Engine to get all raw data
    logger.info("Initializing FilterEngine to fetch KPI columns...")
    engine = FilterEngine(config_path="config/screener_config.yaml", db_path=db_path)
    df_engine = engine.df

    # Ensure all ranked metrics are present in the final export, even if not in KPI_COLUMNS
    export_cols = KPI_COLUMNS.copy()
    if "company_id" in export_cols:
        export_cols.remove("company_id")  # Handled manually

    for m in RANKED_METRICS:
        if m not in export_cols:
            export_cols.append(m)

    # 2. Query DB for mappings, percentiles, and company names
    with sqlite3.connect(db_path) as conn:
        pg_df = pd.read_sql_query(
            "SELECT company_id, peer_group_name, is_benchmark FROM peer_groups", conn
        )
        companies_df = pd.read_sql_query(
            "SELECT id AS company_id, company_name FROM companies", conn
        )
        pp_df = pd.read_sql_query(
            "SELECT company_id, peer_group_name, metric, value, percentile_rank, year_mismatch_flag FROM peer_percentiles",
            conn,
        )

    # 3. Read warnings
    warnings_file = Path("output/peer_group_warnings.csv")
    warnings_dict = {}
    if warnings_file.exists():
        warnings_df = pd.read_csv(warnings_file)
        warnings_dict = dict(
            zip(warnings_df["peer_group_name"], warnings_df["warning"])
        )

    # 4. Prepare Excel Workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    groups = pg_df["peer_group_name"].unique()

    for group in sorted(groups):
        logger.info(f"Processing sheet for peer group: {group}")
        ws = wb.create_sheet(title=str(group)[:31])  # Max 31 chars

        # Determine group's subset
        group_companies = pg_df[pg_df["peer_group_name"] == group]

        # Pivot percentiles for this group
        group_pp = pp_df[pp_df["peer_group_name"] == group]

        # If group has no percentiles, skip or handle gracefully
        if group_pp.empty:
            logger.warning(f"No percentile data for {group}. Continuing anyway.")
            df_ranks = pd.DataFrame()
            mismatches = {}
        else:
            df_ranks = group_pp.pivot(
                index="company_id", columns="metric", values="percentile_rank"
            )
            mismatches = (
                group_pp.drop_duplicates("company_id")
                .set_index("company_id")["year_mismatch_flag"]
                .to_dict()
            )

        # Build values DataFrame
        # Merge engine data with company names and benchmark flag
        df_group_vals = group_companies.merge(df_engine, on="company_id", how="left")
        df_group_vals = df_group_vals.merge(companies_df, on="company_id", how="left")

        # Sort so benchmark is at top, then alphabetical
        df_group_vals = df_group_vals.sort_values(
            by=["is_benchmark", "company_id"], ascending=[False, True]
        )

        # Header Row definition
        headers = ["Company ID", "Company Name"] + export_cols
        ws.append(headers)

        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT

        # Warning caveat
        if group in warnings_dict:
            ws.cell(
                row=1, column=len(headers) + 1, value=f"CAVEAT: {warnings_dict[group]}"
            )
            ws.cell(row=1, column=len(headers) + 1).font = Font(
                color="FF0000", bold=True
            )

        # Write data rows
        row_idx = 2
        for _, row in df_group_vals.iterrows():
            c_id = row["company_id"]
            c_name = row["company_name"]
            is_benchmark = row["is_benchmark"] == 1

            # Format Anchor Year if mismatch
            year_val = row.get("year", "")
            if mismatches.get(c_id, 0) == 1:
                year_val = f"{year_val}*"

            # Prepare row data
            row_data = [c_id, c_name]
            for col in export_cols:
                if col == "year":
                    row_data.append(year_val)
                else:
                    val = row.get(col, pd.NA)
                    row_data.append(val if pd.notna(val) else None)

            ws.append(row_data)

            # Style the row
            for col_i, col_name in enumerate(export_cols):
                cell = ws.cell(row=row_idx, column=col_i + 3)  # Offset for c_id, c_name

                # Numeric formatting
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"

                # Apply conditional color if it's a ranked metric
                if col_name in RANKED_METRICS and c_id in df_ranks.index:
                    rank = df_ranks.at[c_id, col_name]
                    fill = get_percentile_fill(rank)
                    if fill:
                        cell.fill = fill

            # Apply benchmark gold highlight
            if is_benchmark:
                ws.cell(row=row_idx, column=1).fill = GOLD_FILL
                ws.cell(row=row_idx, column=2).fill = GOLD_FILL

            row_idx += 1

        # Summary Row (Median)
        summary_row = ["Summary (Median)", ""]
        numeric_cols = df_group_vals[export_cols].select_dtypes(include=["number"])
        medians = numeric_cols.median()

        for col in export_cols:
            if col in medians:
                summary_row.append(medians[col])
            else:
                summary_row.append(None)

        ws.append(summary_row)
        summary_row_idx = row_idx

        # Style summary row
        for col_idx in range(1, len(summary_row) + 1):
            cell = ws.cell(row=summary_row_idx, column=col_idx)
            cell.font = Font(bold=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

        # Adjust column widths
        for col_idx in range(1, len(headers) + 1):
            letter = get_column_letter(col_idx)
            if col_idx == 2:  # Company Name
                ws.column_dimensions[letter].width = 30
            elif col_idx == 1:
                ws.column_dimensions[letter].width = 15
            else:
                ws.column_dimensions[letter].width = 18

    # Ensure output dir exists
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    logger.info(f"Successfully generated {out_path}")


if __name__ == "__main__":
    generate_peer_comparison()
