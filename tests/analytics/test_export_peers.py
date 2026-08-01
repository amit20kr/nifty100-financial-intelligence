"""
tests/analytics/test_export_peers.py
====================================
Day 20 — Tests for the Peer Comparison Excel Export.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from src.analytics.export_peers import generate_peer_comparison


class TestExportPeers:
    def test_export_generates_correct_sheets(self, tmp_path):
        """
        Integration test verifying the Excel output structure.
        Ensures 11 sheets are created and Summary row exists.
        """
        db_path = Path("db/nifty100.db")
        if not db_path.exists():
            pytest.skip("Integration database db/nifty100.db not found.")

        out_path = tmp_path / "output/peer_comparison.xlsx"

        # Act
        generate_peer_comparison(db_path=db_path, out_path=out_path)

        # Assert File Exists
        assert out_path.exists(), "Excel file must be generated."

        # Open workbook
        wb = openpyxl.load_workbook(out_path)

        # Assert exactly 11 sheets
        # Count might be slightly less if some groups have 0 percentiles, but we should have 11 peer groups
        assert len(wb.sheetnames) == 11, f"Expected 11 sheets, got {len(wb.sheetnames)}"

        # Check first sheet for Summary row
        ws = wb.worksheets[0]

        # Find the last row
        last_row = ws.max_row
        summary_cell = ws.cell(row=last_row, column=1).value
        assert (
            summary_cell == "Summary (Median)"
        ), "The last row must be the Summary (Median) row."

        # Check that headers have 22 columns (Company ID + Name + 20 KPIs)
        # Actually it's 20 KPIs from export.py + 2 missing ones = 22, + 2 id/name = 24.
        header_row = [
            ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)
        ]
        assert (
            "company_id" not in header_row
        ), "company_id should be titled 'Company ID'"
        assert "Company ID" in header_row
        assert "Company Name" in header_row
        assert "debt_to_equity" in header_row
